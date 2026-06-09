from django.shortcuts import render,redirect,get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from .forms import *
from openpyxl import load_workbook, Workbook
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages

# Create your views here.

def ajout_etudiant(request):
    form = EtudiantForm(
        request.POST or None,
        request.FILES or None,
    )

    if form.is_valid():
        form.save()
        return redirect('/')

    return render(request,'app/form.html', {'form':form, 'titre': 'Ajouter un étudiant'})

def ajout_enseignant(request):
    form = EnseignantForm(
        request.POST or None,
    )

    if form.is_valid():
        form.save()
        return redirect('/')

    return render(request, 'app/form.html', {'form':form, 'titre': 'Ajouter un enseignant'})

def ajout_ue(request):
    form = UEForm(
        request.POST or None,
    )
    if form.is_valid():
        form.save()
        return redirect('/app')
    return render(request, 'app/form.html', {'form':form, 'titre': 'Ajouter un ue'})

def ajout_ressource(request):
    prof_id = request.GET.get('prof_id')
    form = RessourceForm(
        request.POST or None,
    )
    if form.is_valid():
        form.save()
        if prof_id:
            return redirect('enseignant', id=prof_id)
        return redirect('/')
    return render(request, 'app/form.html', {'form':form, 'titre': 'Ajout un ressource', 'prof_id': prof_id})

def ajout_examens(request):
    prof_id = request.GET.get('prof_id')
    form = ExamenForm(
        request.POST or None,
    )
    if form.is_valid():
        form.save()
        if prof_id:
            return redirect('enseignant', id=prof_id)
        return redirect('/')
    return render(request, 'app/form.html', {'form':form, 'titre': 'Ajout un examen', 'prof_id': prof_id})

def ajout_notes(request):
    prof_id = request.GET.get('prof_id')
    form = NoteForm(
        request.POST or None,
    )
    if form.is_valid():
        form.save()
        if prof_id:
            return redirect('enseignant', id=prof_id)
        return redirect('/')
    return render(request, 'app/form.html', {'form':form, 'titre': 'Ajout un note', 'prof_id': prof_id})

def modifier_etudiant(request, id):

    etudiant = get_object_or_404(
        Etudiant,
        id=id
    )

    form = EtudiantForm(
        request.POST or None,
        request.FILES or None,
        instance=etudiant
    )

    if form.is_valid():
        form.save()
        return redirect('/')

    return render(request, 'form.html', {
        'form': form,
        'titre': 'Modifier étudiant'
    })


def login_etudiant(request):

    form = LoginEtudiantForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            numero = form.cleaned_data[
                'numero_etudiant'
            ]
            try:
                etudiant = Etudiant.objects.get(
                    numero_etudiant=numero
                )
                return redirect(
                    'profil_etudiant',
                    id=etudiant.id
                )
            except Etudiant.DoesNotExist:
                form.add_error(
                    None,
                    "Étudiant introuvable"
                )
    return render(request, 'app/login.html', {
        'form': form
    })

def profil_etudiant(request, id):

    etudiant = get_object_or_404(
        Etudiant,
        id=id
    )

    # Récupérer toutes les notes de l'étudiant
    notes = Note.objects.filter(etudiant=etudiant).select_related('examen__ressource')
    notes = notes.prefetch_related('examen__ressource__ues')
    
    # Organiser les notes par UE > Ressource
    notes_par_ue = {}
    
    for note in notes:
        ressource = note.examen.ressource
        
        # Boucler à travers les UE liées à la ressource
        for ue in ressource.ues.all():
            if ue not in notes_par_ue:
                notes_par_ue[ue] = {}
            
            if ressource not in notes_par_ue[ue]:
                notes_par_ue[ue][ressource] = []
            
            notes_par_ue[ue][ressource].append(note)

    return render(request, 'app/notes.html', {
        'etudiant': etudiant,
        'notes_par_ue': notes_par_ue
    })

def login_enseignant(request):
    enseignants = Enseignants.objects.all()
    
    return render(request, 'app/login_enseignant.html', {
        'enseignants': enseignants
    })

def profil_enseignant(request, id):
    enseignant = get_object_or_404(
        Enseignants,
        id=id
    )

    return render(request, 'app/enseignant.html', {
        'enseignant': enseignant
    })


def import_notes_excel(request):
    """
    Import des notes via un fichier Excel
    Format attendu: NOM, PRENOM, NOTE1, NOTE2, ...
    """
    prof_id = request.GET.get('prof_id')
    
    if request.method == 'POST':
        form = ImportNotesForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                fichier = request.FILES['fichier_excel']
                examen = form.cleaned_data['examen']
                
                # Lire le fichier Excel
                file_data = fichier.read()
                workbook = load_workbook(io.BytesIO(file_data))
                worksheet = workbook.active
                
                notes_creees = 0
                notes_echouees = 0
                erreurs = []
                
                # Traiter chaque ligne
                for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Ignorer la première ligne si c'est l'en-tête
                        if row[0] and row[0].upper() in ['NOM', 'NAME']:
                            continue
                    
                    try:
                        if not row or not row[0]:  # Ignorer les lignes vides
                            continue
                        
                        nom = row[0].strip() if row[0] else None
                        prenom = row[1].strip() if len(row) > 1 and row[1] else None
                        note_valeur = row[2] if len(row) > 2 else None
                        
                        if not nom or not prenom or note_valeur is None:
                            erreurs.append(f"Ligne {row_idx}: Données incomplètes")
                            notes_echouees += 1
                            continue
                        
                        # Convertir la note en float
                        try:
                            note_valeur = float(note_valeur)
                        except (ValueError, TypeError):
                            erreurs.append(f"Ligne {row_idx}: Note invalide pour {nom} {prenom}")
                            notes_echouees += 1
                            continue
                        
                        # Trouver l'étudiant
                        try:
                            etudiant = Etudiant.objects.get(
                                nom__iexact=nom,
                                prenom__iexact=prenom
                            )
                        except Etudiant.DoesNotExist:
                            erreurs.append(f"Ligne {row_idx}: Étudiant {nom} {prenom} non trouvé")
                            notes_echouees += 1
                            continue
                        except Etudiant.MultipleObjectsReturned:
                            erreurs.append(f"Ligne {row_idx}: Plusieurs étudiants trouvés pour {nom} {prenom}")
                            notes_echouees += 1
                            continue
                        
                        # Créer ou mettre à jour la note
                        note_obj, created = Note.objects.update_or_create(
                            examen=examen,
                            etudiant=etudiant,
                            defaults={'note': note_valeur}
                        )
                        
                        notes_creees += 1
                    
                    except Exception as e:
                        erreurs.append(f"Ligne {row_idx}: Erreur - {str(e)}")
                        notes_echouees += 1
                
                # Afficher les résultats
                messages.success(request, f"✓ {notes_creees} note(s) importée(s) avec succès!")
                if erreurs:
                    for erreur in erreurs[:10]:  # Afficher max 10 erreurs
                        messages.warning(request, f"⚠ {erreur}")
                    if len(erreurs) > 10:
                        messages.warning(request, f"... et {len(erreurs) - 10} autre(s) erreur(s)")
                
                if prof_id:
                    return redirect('enseignant', id=prof_id)
                return redirect('/')
            
            except Exception as e:
                messages.error(request, f"Erreur lors du traitement du fichier: {str(e)}")
    else:
        form = ImportNotesForm()
    
    return render(request, 'app/import_notes.html', {
        'form': form,
        'prof_id': prof_id
    })


def import_etudiants_excel(request):
    """
    Import des étudiants via un fichier Excel
    Format attendu: NUMERO_ETUDIANT, NOM, PRENOM, GROUPE, EMAIL
    """
    if request.method == 'POST':
        form = ImportEtudiantsForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                fichier = request.FILES['fichier_excel']
                
                # Lire le fichier Excel
                file_data = fichier.read()
                workbook = load_workbook(io.BytesIO(file_data))
                worksheet = workbook.active
                
                etudiants_crees = 0
                etudiants_updates = 0
                etudiants_echoues = 0
                erreurs = []
                
                # Traiter chaque ligne
                for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                    if row_idx == 1:  # Ignorer la première ligne si c'est l'en-tête
                        if row[0] and row[0].upper() in ['NUMERO_ETUDIANT', 'NUMERO', 'NUMBER']:
                            continue
                    
                    try:
                        if not row or not row[0]:  # Ignorer les lignes vides
                            continue
                        
                        numero_etudiant = str(row[0]).strip() if row[0] else None
                        nom = row[1].strip() if len(row) > 1 and row[1] else None
                        prenom = row[2].strip() if len(row) > 2 and row[2] else None
                        groupe = row[3].strip() if len(row) > 3 and row[3] else None
                        email = row[4].strip() if len(row) > 4 and row[4] else None
                        
                        if not numero_etudiant or not nom or not prenom or not groupe or not email:
                            erreurs.append(f"Ligne {row_idx}: Données incomplètes (tous les champs sont obligatoires)")
                            etudiants_echoues += 1
                            continue
                        
                        # Vérifier que l'email est valide
                        if '@' not in email:
                            erreurs.append(f"Ligne {row_idx}: Email invalide pour {nom} {prenom}")
                            etudiants_echoues += 1
                            continue
                        
                        # Créer ou mettre à jour l'étudiant
                        etudiant, created = Etudiant.objects.update_or_create(
                            numero_etudiant=numero_etudiant,
                            defaults={
                                'nom': nom,
                                'prenom': prenom,
                                'groupe': groupe,
                                'email': email
                            }
                        )
                        
                        if created:
                            etudiants_crees += 1
                        else:
                            etudiants_updates += 1
                    
                    except Exception as e:
                        erreurs.append(f"Ligne {row_idx}: Erreur - {str(e)}")
                        etudiants_echoues += 1
                
                # Afficher les résultats
                if etudiants_crees > 0:
                    messages.success(request, f"✓ {etudiants_crees} étudiant(s) créé(s) avec succès!")
                if etudiants_updates > 0:
                    messages.success(request, f"✓ {etudiants_updates} étudiant(s) mis à jour!")
                if erreurs:
                    for erreur in erreurs[:10]:  # Afficher max 10 erreurs
                        messages.warning(request, f"⚠ {erreur}")
                    if len(erreurs) > 10:
                        messages.warning(request, f"... et {len(erreurs) - 10} autre(s) erreur(s)")
                
                return redirect('import_etudiants')
            
            except Exception as e:
                messages.error(request, f"Erreur lors du traitement du fichier: {str(e)}")
    else:
        form = ImportEtudiantsForm()
    
    return render(request, 'app/import_etudiants.html', {
        'form': form
    })


def export_notes_etudiant_pdf(request, id):
    """
    Exporte les notes d'un étudiant en PDF
    """
    etudiant = get_object_or_404(Etudiant, id=id)
    
    # Récupérer les notes de l'étudiant
    notes = Note.objects.filter(etudiant=etudiant).select_related('examen__ressource')
    notes = notes.prefetch_related('examen__ressource__ues')
    
    # Organiser les notes par UE > Ressource
    notes_par_ue = {}
    for note in notes:
        ressource = note.examen.ressource
        for ue in ressource.ues.all():
            if ue not in notes_par_ue:
                notes_par_ue[ue] = {}
            if ressource not in notes_par_ue[ue]:
                notes_par_ue[ue][ressource] = []
            notes_par_ue[ue][ressource].append(note)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="notes_{etudiant.numero_etudiant}.pdf"'
    
    # Créer le document PDF
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    ue_style = ParagraphStyle(
        'UETitle',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#0066cc'),
        spaceAfter=10,
        spaceBefore=10
    )
    ressource_style = ParagraphStyle(
        'RessourceTitle',
        parent=styles['Heading3'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),
        spaceAfter=5
    )
    
    # Éléments du document
    story = []
    
    # En-tête
    title = f"RELEVÉ DE NOTES - {etudiant.prenom} {etudiant.nom}"
    story.append(Paragraph(title, title_style))
    
    # Infos étudiant
    info_text = f"<b>N° Étudiant:</b> {etudiant.numero_etudiant} | <b>Groupe:</b> {etudiant.groupe} | <b>Email:</b> {etudiant.email}"
    story.append(Paragraph(info_text, styles['Normal']))
    
    info_text2 = f"<b>Généré le:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    story.append(Paragraph(info_text2, styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Tableau des notes
    if notes_par_ue:
        for ue, ressources in notes_par_ue.items():
            # Titre de l'UE
            ue_title = f"{ue.code} - {ue.nom} (S{ue.semestre}, {ue.credits_ects} ECTS)"
            story.append(Paragraph(ue_title, ue_style))
            
            # Pour chaque ressource
            for ressource, notes_list in ressources.items():
                ressource_title = f"{ressource.code} - {ressource.nom} (Coef. {ressource.coefficient})"
                story.append(Paragraph(ressource_title, ressource_style))
                
                # Tableau des notes
                data = [['Examen', 'Date', 'Coefficient', 'Note']]
                
                for note in notes_list:
                    data.append([
                        note.examen.title,
                        note.examen.date.strftime('%d/%m/%Y'),
                        str(note.examen.coefficient),
                        f"{note.note}/20"
                    ])
                
                table = Table(data, colWidths=[6*cm, 3*cm, 2*cm, 2*cm])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
                ]))
                
                story.append(table)
                story.append(Spacer(1, 0.3*cm))
    else:
        story.append(Paragraph("Aucune note disponible", styles['Normal']))
    
    # Générer le PDF
    doc.build(story)
    return response


def export_notes_prof_pdf(request, id):
    """
    Exporte toutes les notes gérées par un prof en PDF
    """
    enseignant = get_object_or_404(Enseignants, id=id)
    
    # Récupérer les examens du prof (via les ressources)
    # Pour simplifier, on prend toutes les notes liées aux examens
    notes = Note.objects.select_related('examen__ressource', 'etudiant')
    notes = notes.prefetch_related('examen__ressource__ues')
    
    # Organiser par UE > Ressource > Examen
    data_par_ue = {}
    for note in notes:
        ressource = note.examen.ressource
        for ue in ressource.ues.all():
            if ue not in data_par_ue:
                data_par_ue[ue] = {}
            if ressource not in data_par_ue[ue]:
                data_par_ue[ue][ressource] = {}
            if note.examen not in data_par_ue[ue][ressource]:
                data_par_ue[ue][ressource][note.examen] = []
            data_par_ue[ue][ressource][note.examen].append(note)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="notes_prof_{enseignant.id}.pdf"'
    
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=10,
        alignment=TA_CENTER
    )
    ue_style = ParagraphStyle(
        'UETitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#0066cc'),
        spaceAfter=8,
        spaceBefore=8
    )
    ressource_style = ParagraphStyle(
        'RessourceTitle',
        parent=styles['Heading3'],
        fontSize=9,
        textColor=colors.HexColor('#333333'),
        spaceAfter=3
    )
    
    story = []
    
    # En-tête
    title = f"RELEVÉ DES NOTES - {enseignant.prenom} {enseignant.nom}"
    story.append(Paragraph(title, title_style))
    
    info_text = f"<b>Généré le:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    story.append(Paragraph(info_text, styles['Normal']))
    story.append(Spacer(1, 0.3*cm))
    
    # Contenu
    if data_par_ue:
        for ue in sorted(data_par_ue.keys(), key=lambda x: (x.semestre, x.code)):
            ressources = data_par_ue[ue]
            ue_title = f"{ue.code} - {ue.nom} (S{ue.semestre})"
            story.append(Paragraph(ue_title, ue_style))
            
            for ressource in sorted(ressources.keys(), key=lambda x: x.code):
                examens = ressources[ressource]
                ressource_title = f"{ressource.code} - {ressource.nom} (Coef. {ressource.coefficient})"
                story.append(Paragraph(ressource_title, ressource_style))
                
                for examen in sorted(examens.keys(), key=lambda x: x.title):
                    notes_list = examens[examen]
                    
                    # Tableau
                    data = [[f'Examen: {examen.title}', 'Étudiant', 'Note']]
                    
                    for note in sorted(notes_list, key=lambda x: (x.etudiant.nom, x.etudiant.prenom)):
                        data.append([
                            examen.date.strftime('%d/%m/%Y'),
                            f"{note.etudiant.prenom} {note.etudiant.nom}",
                            f"{note.note}/20"
                        ])
                    
                    table = Table(data, colWidths=[4*cm, 5*cm, 2*cm])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066cc')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('FONTSIZE', (0, 1), (-1, -1), 7),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
                    ]))
                    
                    story.append(table)
                    story.append(Spacer(1, 0.2*cm))
                story.append(Spacer(1, 0.1*cm))
            story.append(Spacer(1, 0.2*cm))
    else:
        story.append(Paragraph("Aucune note disponible", styles['Normal']))
    
    doc.build(story)
    return response
# ===== LOGIN ENSEIGNANT =====
def login_enseignant(request):
    form = LoginEnseignantForm(request, data=request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)

            if user is not None:
                # Vérifier qu'il est bien lié à un Enseignant
                if hasattr(user, 'enseignant'):
                    login(request, user)
                    return redirect('enseignant', id=user.enseignant.id)
                else:
                    form.add_error(None, "Ce compte n'est pas un compte enseignant.")
            else:
                form.add_error(None, "Identifiants incorrects.")

    return render(request, 'app/login_enseignant.html', {'form': form})


# ===== LOGOUT =====
def logout_enseignant(request):
    logout(request)
    messages.success(request, "Vous avez été déconnecté.")
    return redirect('login_etudiant')


# ===== PROFIL ENSEIGNANT (protégé) =====
@login_required(login_url='login_enseignant')
def profil_enseignant(request, id):
    enseignant = get_object_or_404(Enseignants, id=id)

    # Sécurité : un prof ne peut voir que son propre profil
    if not hasattr(request.user, 'enseignant') or request.user.enseignant.id != enseignant.id:
        messages.error(request, "Accès refusé.")
        return redirect('login_enseignant')

    return render(request, 'app/enseignant.html', {'enseignant': enseignant})


# ===== AJOUT ENSEIGNANT (avec création compte User) =====
def ajout_enseignant(request):
    form = EnseignantCreationForm(request.POST or None)

    if form.is_valid():
        # Créer l'utilisateur Django
        user = User.objects.create_user(
            username=form.cleaned_data['username'],
            password=form.cleaned_data['password']
        )
        # Créer l'enseignant lié
        enseignant = form.save(commit=False)
        enseignant.user = user
        enseignant.save()

        messages.success(request, f"Enseignant {enseignant.prenom} {enseignant.nom} créé !")
        return redirect('login_enseignant')

    return render(request, 'app/form.html', {
        'form': form,
        'titre': 'Ajouter un enseignant'
    })